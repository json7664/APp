import pandas as pd
import numpy as np
import os
import win32com.client
from openpyxl import load_workbook
import time
import colorama

start = time.time()
curr_working_dir = os.getcwd()

pathTo_MACRO = curr_working_dir + "\\Compare.txt"
pathTo_XLSM = curr_working_dir + "\\Collected_data.xlsm"

TS_database = open(curr_working_dir + "\\etc\\TS.txt", 'r').read()
VS_database = open(curr_working_dir + "\\etc\\VS.txt", 'r').read()
MDAC_database = open(curr_working_dir + "\\etc\\MDAC.txt", 'r').read()
BG_database = open(curr_working_dir + '\\etc\\BG.txt', 'r').read()
LO_database = open(curr_working_dir + '\\etc\\LO.txt', 'r').read()
PI_database = open(curr_working_dir + '\\etc\\PI.txt', 'r').read()
OTP_database = open(curr_working_dir + '\\etc\\OTP.txt', 'r').read()
NOL_database = open(curr_working_dir + '\\etc\\NOL.txt', 'r').read()
Speedmon_database = open(curr_working_dir + '\\etc\\SPM.txt', 'r').read()
UVLO_database = open(curr_working_dir + '\\etc\\UVLO.txt', 'r').read()
IS_database = open(curr_working_dir + '\\etc\\IS.txt', 'r').read()
IIL_database = open(curr_working_dir + '\\etc\\IIL.txt', 'r').read()
IIH_database = open(curr_working_dir + '\\etc\\IIH.txt', 'r').read()
RDS_database = open(curr_working_dir + '\\etc\\RDS.txt', 'r').read()
VOL_database = open(curr_working_dir + '\\etc\\VOL.txt', 'r').read()
VOH_database = open(curr_working_dir + '\\etc\\VOH.txt', 'r').read()
IOZ_database = open(curr_working_dir + '\\etc\\IOZ.txt', 'r').read()
PS_database = open(curr_working_dir + '\\etc\\PS.txt', 'r').read()
SP_database = open(curr_working_dir + '\\etc\\SP.txt', 'r').read()
PRDIN_database = open(curr_working_dir + '\\etc\\PRDIN.txt', 'r').read()

User_input = input("Number of dataset(s) to compute: ")
path_lst = []
sheet_lst = []
print("Example path: C:\\Users\\XXXX\\XXXXX\\XXXXX\\XXXXXXX.csv")
for i in range(int(User_input)):
    Path_input = input(f"Enter the path of dataset {i + 1}: ")
    Sheet_input = input(f"Enter the Sheet name {i + 1}: ")
    path_lst.append(Path_input)
    sheet_lst.append(Sheet_input)

file_dict = {}
for file in path_lst:
    key = file
    df = pd.DataFrame(pd.read_csv(file))

    file_dict[key] = df

counter = 0
writer = pd.ExcelWriter('Combined.xlsx', engine='xlsxwriter')
for path in file_dict:
    df = file_dict[path]
    lst = []
    for name in df["TESTNAME"]:
        if "Cont" in str(name):
            lst.append('CT')
        elif "Scan" in str(name):
            lst.append("SC")
        elif "IDDQ" in str(name):
            lst.append("ID_VD")
        elif name in TS_database:
            lst.append("TS")
        elif name in VS_database:
            lst.append("VS")
        elif name in MDAC_database:
            lst.append("MDAC")
        elif name in BG_database:
            lst.append("BG")
        elif name in LO_database:
            lst.append("LO")
        elif name in PI_database:
            lst.append("PI")
        elif name in OTP_database:
            lst.append("OTP")
        elif name in NOL_database:
            lst.append("NOL")
        elif name in Speedmon_database:
            lst.append("SPM")
        elif name in UVLO_database:
            lst.append("UVLO")
        elif name in IS_database:
            lst.append("IS")
        elif name in IIL_database:
            lst.append("IIL")
        elif name in IIH_database:
            lst.append("IIH")
        elif name in RDS_database:
            lst.append("RDS")
        elif name in VOL_database:
            lst.append("VOL")
        elif name in VOH_database:
            lst.append("VOH")
        elif name in IOZ_database:
            lst.append("IOZ")
        elif name in PS_database:
            lst.append("PS")
        elif name in SP_database:
            lst.append("SP")
        elif name in PRDIN_database:
            lst.append("PRDIN")
        else:
            lst.append('Nan')

    for i in range(9):
        lst[i] = 'OTH'

    lst[0] = ""

    df['Abbreviations'] = np.array(lst)

    df.to_excel(writer, sheet_name=sheet_lst[counter])
    if counter >= int(User_input):
        counter -= 1
    else:
        counter += 1
writer.save()

# print('\n\x1b[4;30;41m' + "To Stop, enter STOP in the input" + '\x1b[0m')
src_lst = []
search = {}
print("To Stop, enter STOP in the input")
while True:
    TST_lst = []
    BRFlag = False
    Usr_sheet_input = input("Sheet(s) to be searched from: ")
    if Usr_sheet_input == "STOP" or Usr_sheet_input == "stop":
        break
    src_lst.append(Usr_sheet_input)
    for i in src_lst:
        if i not in sheet_lst:
            print("Enter sheet name that is in Combined.xlsx")
            src_lst.remove(i)
    while not BRFlag:
        TST_input = input("Enter test-block abbreviation(s) to be searched: ")
        if TST_input == "STOP" or TST_input == "stop":
            BRFlag = True
        else:
            TST_lst.append(TST_input)
    search[Usr_sheet_input] = TST_lst

# TODO merge the data and make isolated test block sheets

writer2 = pd.ExcelWriter("Collected_data.xlsx", engine='xlsxwriter')

for sheetName in search:
    sN = sheetName
    lst_ts = search[sN]
    df = pd.read_excel(curr_working_dir + '\\Combined.xlsx', sheet_name=sN)
    for i in lst_ts:
        df1 = df.loc[df["Abbreviations"] == i]
        # collected_df.to_excel(writer2, sheet_name="{TestName}_{Block}".format(TestName=sN, Block=i))
        # df1.to_csv(r"C:\Users\ShendeJ\Desktop\CMD_script\{Name}_{sheet}_BLOCK.csv".format(Name=sN, sheet=i))
        del df1["Unnamed: 0"]
        df1.to_excel(writer2, sheet_name="{Name}_{Block}".format(Name=sN, Block=i))

writer2.save()

# writer3 = pd.ExcelWriter("C2_data.xlsx", engine='xlsxwriter')
# combined_df = pd.DataFrame(pd.read_excel(r"C:\Users\ShendeJ\Desktop\CMD_script\Combined.xlsx", sheet_name=None))
# collected_df = pd.DataFrame(pd.read_excel(r'C:\Users\ShendeJ\Desktop\CMD_script\Collected_data.xlsx', \
# sheet_name=None))
# df2 = pd.concat([combined_df, collected_df], ignore_index=True)
# df2.to_excel(r"C:\Users\ShendeJ\Desktop\CMD_script\C2_data.xlsx")
#  writer3.save()


fn = curr_working_dir + '\\Collected_data.xlsx'
book = load_workbook(fn)
# Get McK dataframe
df_mas = pd.read_csv(curr_working_dir + '\\Master.csv')
# pd.read_excel returns a 'dict', where key is sheet name and value is the dataframe
# get only dataframe, so get value of key 'McK'
writer3 = pd.ExcelWriter(fn, engine='openpyxl')
writer3.book = book
writer3.book.create_sheet('Dump')
df_mas.to_excel(writer3, sheet_name='Master_Sheet')
writer3.save()

# TODO add injection script
# directory = "Output"
# parent_dir = curr_working_dir
# path = os.path.join(parent_dirm)

excel = win32com.client.Dispatch('Excel.Application')
wb = excel.Workbooks.Open(curr_working_dir + "\\Collected_data.xlsx")
excel.DisplayAlerts = False
wb.DoNotPromptForConvert = True
wb.CheckCompatibility = False
wb.SaveAs(curr_working_dir + "\\Collected_data.xlsm", FileFormat=52)
excel.Quit()
os.system('TASKKILL /F /IM excel.exe')

# os.remove('Collected_data.xlsx')

MACRO_name = 'Compare'

with open(pathTo_MACRO, "r") as my_file:
    print('reading macro into string from: ' + str(my_file))
    macro = my_file.read()

excel1 = win32com.client.Dispatch("Excel.Application")
workbook = excel1.Workbooks.Open(Filename=pathTo_XLSM)

# insert the macro-string into the Excel file
excelModule = workbook.VBProject.VBComponents.Add(1)
excelModule.CodeModule.AddFromString(macro)

# save the workbook and close
excel1.Workbooks(1).Close(SaveChanges=1)
excel1.Application.Quit()

# garbage collection
del excel
print(colorama.Fore.BLUE + 'Output data (Collected_data.xlsm) will be in the same program directory')
print(colorama.Fore.YELLOW + 'Warning! Please do not open Collected_data.xlsm file at this moment')

xl_ = win32com.client.Dispatch("Excel.Application")  # instantiate excel app
# script_Name = 'Yos_Ran_McK_Limits.xlsm!Module1.TestComp'
# path = 'C:\\Users\\ShendeJ\\PycharmProjects\\pythonProject\\Yos_Ran_McK_Limits.xlsm'
print(colorama.Fore.RED + 'Script Running!')
wb = xl_.Workbooks.Open(curr_working_dir + "\\Collected_data.xlsm")
xl_.Application.Run('Collected_data.xlsm!Module1.Compare')
wb.Save()
xl_.Application.Quit()

print(colorama.Fore.GREEN + 'Successfully compared Limits!')

end = time.time()

print("{Mins} mins".format(Mins=(end - start) / 60))
